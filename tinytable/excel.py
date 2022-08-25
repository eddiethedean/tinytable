from typing import Mapping, Optional, Union
from os.path import exists

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from openpyxl.chartsheet.chartsheet import Chartsheet

from tinytable.functional.utils import combine_names_rows
from tinytable.functional.rows import itertuples



Sheet = Union[Worksheet, ReadOnlyWorksheet, Chartsheet]
WorkSheet = Union[Worksheet, ReadOnlyWorksheet]


class WorkBook:
    def __init__(self, path: str) -> None:
        self.path = path

    def __enter__(self):
        self.wb = load_workbook(self.path)
        return self

    def __exit__(self, exc_type, exc_value, exc_traceback):
        self.wb.close()

    @property
    def active(self) -> Worksheet:
        return self.wb.active

    def __getitem__(self, key: str) -> Sheet:
        return self.wb[key]


def read_excel_file(path: str, sheet_name: Optional[str] = None) -> dict:
    """
    Reads a table object from given excel file path.
    """
    column_names = []
    rows = []
    first = True
    with WorkBook(path) as wb:
        ws = wb.active if sheet_name is None else wb[sheet_name]
        if isinstance(ws, Chartsheet):
            raise TypeError('Chartsheet has no values to read into table.')
        for row in ws.values:
            if first:
                column_names = row
                first = False
            else:
                rows.append(row)

    return combine_names_rows(column_names, rows)


def data_to_excel_file(data: Mapping, path: str, sheet_name: Optional[str] = None) -> None:
    """
    Write data to excel file.
    Path needs to end with file name then .xlsx
    Creates new xlsx file if path file does not exist.
    Adds new worksheet named sheet_name if the file exists.
    Overides worksheet sheet_name if it already exists.
    """
    if exists(path):
        wb = load_workbook(path)
        if sheet_name is None:
            sheet_name = wb.active.title
    else:
        wb = Workbook()
        if sheet_name is None:
            sheet_name = 'Sheet1'
        
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.delete_rows(1, ws.max_row+1)
    else:
        ws = wb.create_sheet(sheet_name)
    
    # add column names
    ws.append(list(data.keys()))
    # add rows data
    for row in itertuples(data):
        ws.append(row)
    wb.save(path)
    